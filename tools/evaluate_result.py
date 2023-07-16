# -*- coding: utf-8 -*-

# @Time:      2023/01/22 15:40
# @Author:    qiao
# @Email:     rukunqiao@outlook.com
# @File:      evaluate_result.py
# @Software:  PyCharm
# @Description:
#   This file is used for disparity map evaluation.

# - Package Imports - #
from configparser import ConfigParser
from pathlib import Path
import numpy as np
import torch
import openpyxl
from openpyxl.styles import Alignment
from tqdm import tqdm

import utils.pointerlib as plb


# - Coding Part - #
class Evaluator:
    def __init__(self, workbook, append_flag):
        self.workbook = openpyxl.load_workbook(str(workbook))
        self.workbook_path = workbook
        self.append_flag = append_flag

    def process_dataset(self, scene_name, mask_flag=False):
        work_sheet = self.workbook[scene_name]
        data_path = Path(work_sheet['B1'].value)
        out_path = Path(work_sheet['B2'].value)

        # load parameters
        config = ConfigParser()
        config.read(str(data_path / 'config.ini'), encoding='utf-8')
        seq_num = int(config['Data']['total_sequence'])
        frm_len = plb.str2tuple(config['Data']['frm_len'], int)
        if len(frm_len) == 1:
            frm_len = frm_len * seq_num
        wid, hei = plb.str2tuple(config['Data']['img_size'], item_type=int)

        # Get all exp_tags
        start_row = 5
        exp_set = sorted([x for x in out_path.glob('*') if x.is_dir()])
        if self.append_flag:
            exp_set_exists = [work_sheet.cell(i, 1).value for i in range(start_row, work_sheet.max_row + 1)]
            exp_set_new = [x for x in exp_set if x.name not in exp_set_exists]
            exp_set = exp_set_new
            start_row = work_sheet.max_row + 1
        exp_num = len(exp_set)

        # Build up cmp_sets
        cmp_sets = []
        for i, exp_path in enumerate(exp_set):
            work_sheet.cell(row=start_row + i, column=1).value = exp_path.name
            sub_folders = [x for x in (exp_path / 'output' / data_path.name).glob('*') if x.is_dir()]
            res_path = sorted(sub_folders)[-1]
            work_sheet.cell(row=start_row + i, column=2).value = str(res_path)
            cmp_sets.append(self._build_up_cmp_set(data_path, res_path, frm_len, mask_flag))
            
        # Evaluate: column = 3,4,5,6
        prog_bar = tqdm(total=sum([len(x) for x in cmp_sets]))
        for i in range(exp_num):
            prog_bar.set_description(f'{scene_name}[{exp_set[i].name}]')
            row_idx = start_row + i

            # Evaluate result and fill results
            res = self._evaluate_exp_outs(cmp_sets[i], prog_bar)
            work_sheet.cell(row=row_idx, column=3).value = f'{res[0] * 100.0:.2f}'
            work_sheet.cell(row=row_idx, column=4).value = f'{res[1] * 100.0:.2f}'
            work_sheet.cell(row=row_idx, column=5).value = f'{res[2] * 100.0:.2f}'
            work_sheet.cell(row=row_idx, column=6).value = f'{res[3]:.3f}'

            # Style
            work_sheet.row_dimensions[row_idx].height = 15
            for col_idx in range(1, 7):
                work_sheet.cell(row_idx, col_idx).alignment = Alignment('center', 'center')

        self.workbook.save(self.workbook_path)
        pass

    def _build_up_cmp_set(self, data_path, res_path, frm_len, mask_flag=False):
        cmp_set = []
        
        seq_num = len([x for x in res_path.glob('scene_*') if x.is_dir()])
        for seq_idx in range(seq_num):
            gt_scene_folder = data_path / f'scene_{seq_idx:04}'
            disp_gt_folder = gt_scene_folder / 'disp_gt'
            if not disp_gt_folder.exists():
                disp_gt_folder = gt_scene_folder / 'disp'
            res_scene_folder = res_path / f'scene_{seq_idx:04}'
            disp_res_folder = res_scene_folder / 'disp'

            for frm_idx in range(frm_len[seq_idx]):
                disp_gt_path = disp_gt_folder / f'disp_{frm_idx}.png'
                disp_res_path = disp_res_folder / f'disp_{frm_idx}.png'
                mask_path = gt_scene_folder / 'mask' / f'mask_{frm_idx}.png' if mask_flag else None

                if disp_gt_path.exists() and disp_res_path.exists():
                    cmp_set.append(tuple([disp_gt_path, disp_res_path, mask_path]))

        return cmp_set

    def _evaluate_exp_outs(self, cmp_set, prog_bar=None):
        res_array = []

        for disp_gt_path, disp_res_path, mask_path in cmp_set:

            disp_gt = plb.imload(disp_gt_path, scale=1e2)
            disp_res = plb.imload(disp_res_path, scale=1e2)
            mask = (disp_gt > 0.0).float()  # TODO
            if mask_path is not None:
                mask = plb.imload(mask_path)

            diff = (disp_gt - disp_res)
            diff_vec = diff[mask > 0.0]
            total_num = diff_vec.shape[0]
            err10_num = (torch.abs(diff_vec) > 1.0).float().sum() / total_num
            err20_num = (torch.abs(diff_vec) > 2.0).float().sum() / total_num
            err50_num = (torch.abs(diff_vec) > 5.0).float().sum() / total_num
            avg = torch.abs(diff_vec).sum() / total_num
            res_array.append(np.array([err10_num, err20_num, err50_num, avg]))

            if prog_bar is not None:
                prog_bar.update(1)

        res_array = np.stack(res_array, axis=0)
        res_avg = np.average(res_array, axis=0)
        return res_avg

    def sum_average(self, scene_name):
        src_work_sheet = self.workbook[scene_name]
        dst_work_sheet = self.workbook[f'{scene_name}-Sum']

        # Get all exp_tags & values
        start_row = 5
        raw_results = []
        for row_idx in range(start_row, src_work_sheet.max_row + 1):
            raw_results.append([
                src_work_sheet.cell(row_idx, 1).value,
                [
                    float(src_work_sheet.cell(row_idx, 3).value),
                    float(src_work_sheet.cell(row_idx, 4).value),
                    float(src_work_sheet.cell(row_idx, 5).value),
                    float(src_work_sheet.cell(row_idx, 6).value),
                ]
            ])

        # Group
        grouped_results = {}
        for exp_tag, eval_res in raw_results:
            exp_key = exp_tag
            if '_exp' in exp_tag:
                exp_key = exp_tag.split('_exp')[0]
            if exp_key not in grouped_results:
                grouped_results[exp_key] = []
            grouped_results[exp_key].append(np.array(eval_res))

        # Put into new sheet
        for i, exp_tag in enumerate(grouped_results):
            dst_work_sheet.cell(2 + i, 1).value = exp_tag
            exp_res = np.stack(grouped_results[exp_tag], axis=0)
            exp_avg = np.average(exp_res, axis=0)
            dst_work_sheet.cell(2 + i, 2).value = f'{exp_avg[0]:.2f}'
            dst_work_sheet.cell(2 + i, 3).value = f'{exp_avg[1]:.2f}'
            dst_work_sheet.cell(2 + i, 4).value = f'{exp_avg[2]:.2f}'
            dst_work_sheet.cell(2 + i, 5).value = f'{exp_avg[3]:.3f}'

        self.workbook.save(self.workbook_path)
        pass


def main():
    app = Evaluator(
        workbook='./res/result.xlsx',  # Here is your excel file
        append_flag=False
    )
    app.process_dataset('NonRigidReal', mask_flag=False)  # This is the sheet name.
    app.sum_average('NonRigidReal')  # Results will be saved in sheet 'NonRigidReal-Sum'.


if __name__ == '__main__':
    main()
